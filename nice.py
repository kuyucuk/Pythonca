import urllib.request
import bs4 as bs
import datetime


from openpyxl import Workbook
wb = Workbook()
ws = wb.active

sauce = urllib.request.urlopen('https://www.ubersem.com').read()

soup = bs.BeautifulSoup(sauce,'lxml')


h1data=soup.h1

h1text=h1data.text

ws['A1'] = str(h1text)



ws['B3']=(str(soup.title.text))



data=soup.h2

text=data.text

ws['D5']= str(text)



ws['E7']=(str(soup.h3.text))


ws['G9'] = datetime.datetime.now()


address_box = soup.find('div', attrs={'class':'mk-col-1-4'})
address = address_box.text

ws['I11']= str(address)



ws['K13']= str(soup.find('span', attrs={'id':'fancy-title-8'}).text)



wb.save("yaz.xlsx")

